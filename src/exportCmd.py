# Expects an excel spreadsheet as exported from moodle with additional colums
# A1 (5), A2 (6) for each assignment. The numbers in parens are the maximal points
# of the assignment.

from config import *
from ownLogging import *
from utils import *
import re
from dataclasses import dataclass
import tempfile
import csv
import utils

@dataclass
class PointCol:
    idx: int
    num: int
    maxPoints: int
    def format(self, cfg, points):
        if cfg.lang == 'de':
            return f'Aufgabe {self.num}: {points} von {self.maxPoints} Punkten'
        else:
            return f'Exercise {self.num}: {points} of {self.maxPoints} points'

pointColRe = re.compile(r'A(?P<num>[0-9]+) +\((?P<points>[^)]+)\)')

def parsePointCol(s, idx):
    m = pointColRe.match(s)
    if m:
        num = int(m.group('num'))
        maxPoints = m.group('points')
        return PointCol(idx, num, maxPoints)
    else:
        return None

def fixStudentId(s):
    prefix = 'Teilnehmer/in'
    if s.startswith(prefix):
        return s[len(prefix):]
    else:
        return s

STATUS_COL = 'Status'
ID_COL = 'ID'
RATING_COL = 'Bewertung'
NAME_COL = 'Vollständiger Name'

def loadShpreadsheet(name):
    import openpyxl as exc
    wb = exc.load_workbook(filename=name, data_only=True)
    return wb

def export(cfg):
    f = cfg.spreadsheetPath
    wb = loadShpreadsheet(f)
    sheetnames = wb.sheetnames
    if len(sheetnames) == 0:
        abort(f'Spreadsheet {f} does not define any sheets')
    sheet = wb[sheetnames[0]]
    values = list(sheet.values)
    if len(values) == 0:
        abort(f'Spreadsheet {f} does not contain any data')
    titleCols = [(i, s.strip() if s else None) for (i, s) in enumerate(values[0])]
    statusIdx = None
    idIdx = None
    ratingIdx = None
    pointColIdxs = []
    pointCols = []
    for i, s in titleCols:
        if s == STATUS_COL:
            statusIdx = i
        elif s == ID_COL:
            idIdx = i
        elif s == RATING_COL:
            ratingIdx = i
        elif s:
            c = parsePointCol(s, i)
            if c:
                pointCols.append(c)
                pointColIdxs.append(i)
    pointColIdxs.sort(reverse=True)
    if statusIdx is None:
        abort(f'No column {STATUS_COL} found in {f}')
    if idIdx is None:
        abort(f'No colum {ID_COL} found in {f}')
    if ratingIdx is None:
        abort(f'No colum {RATING_COL} found in {f}')
    # write POINTS.txt file
    pointsTemplate = ''
    if os.path.isfile(cfg.pointsTemplate):
        pointsTemplate = utils.readFile(cfg.pointsTemplate).strip()
        if pointsTemplate:
            pointsTemplate = pointsTemplate + "\n\n"
    for student in values[1:]:
        status = student[statusIdx]
        studentId = student[idIdx]
        if not studentId:
            continue
        studentId = fixStudentId(studentId)
        if status and status.lower() == 'keine abgabe':
            continue
        lines = []
        for c in pointCols:
            points = student[c.idx]
            lines.append(c.format(cfg, points))
        d = findSubmissionDirForId(cfg, studentId)
        if not d:
            warn(f'No submission directory for student with ID {studentId}')
        else:
            f = shell.pjoin(d, cfg.pointsFile)
            writeFile(f, pointsTemplate + '\n'.join(lines))
            verbose(f'Wrote {f}')
    # write .csv file
    ratingCsv = cfg.ratingCsvPath
    with open(ratingCsv, 'w', encoding='utf-8') as csvfile:
        w = csv.writer(csvfile, quoting=csv.QUOTE_MINIMAL)
        for v in values:
            l = list(v)
            # rating = l[ratingIdx]
            # if type(rating) is int:
            #     rating = '%d' % rating
            # elif type(rating) is float:
            #     rating = ('%.1f' % rating).replace('.', ',')
            # l[ratingIdx] = rating
            for i in pointColIdxs:
                # pointColIdxs is sorted in reverse
                l.pop(i)
            w.writerow(l)
    print(f'Wrote {ratingCsv}')
    # write .zip file feedback
    with shell.workingDir(cfg.baseDir):
        allDirs = collectSubmissionDirs(cfg, baseDir='.')
        shell.removeFile(cfg.feedbackZip)
        # shell.run(['zip', '-q', '-r', cfg.feedbackZip] + allDirs + ['-x', '*/TUTOR/*'])
        zipDirs(cfg.feedbackZip, allDirs, excludeDirs=['.stack-work'])
    print(f'Wrote {shell.pjoin(cfg.baseDir, cfg.feedbackZip)}')
