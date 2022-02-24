import abc
import typing
import tempfile
import os
import atexit
from ownLogging import *

#
# Abstract class abstract over certain aspects of excel files
#
# Columns and rows are indexed 0,1,2,..
class Sheet(abc.ABC):
    @abc.abstractmethod
    def getNumberOfRows(self) -> int:
        pass

    @property
    def numberOfRows(self) -> int:
        return self.getNumberOfRows()

    def rowIndices(self, start=0, end=None):
        if end is None:
            end = self.numberOfRows
        return range(start, end)

    @abc.abstractmethod
    def getNumberOfCols(self) -> int:
        pass

    @property
    def numberOfCols(self) -> int:
        return self.getNumberOfCols()

    def colIndices(self, start=0, end=None):
        if end is None:
            end = self.numberOfCols
        return range(start, end)

    @abc.abstractmethod
    def value(self, col: int, row: int):
        pass

    @abc.abstractmethod
    def setValue(self, col: int, row: int, value: any):
        pass

    def row(self, row: int):
        result = []
        for i in range(0, self.numberOfCols):
            result.append(self.value(col=i, row=row))
        return result

    def rows(self) -> typing.Iterable[list]:
        for i in range(self.numberOfRows):
            yield self.row(i)

    @abc.abstractmethod
    def save(self, path):
        pass

    def close(self):
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, value, traceback):
        self.close()

def saveBook(book, path):
    # I experienced problem with corrupt .xlsx files, probably when the python process
    # was interrupted while writing the data.
    with tempfile.NamedTemporaryFile(suffix='.xlsx', dir=os.path.dirname(path)) as fp:
        fp.close()
        book.save(fp.name)
        os.rename(fp.name, path) # Atomic move

# Openpyxl does not support formulae but it also works with Linux (xlwings doesn't)
class OpenpyxlSheet(Sheet):
    def __init__(self, wb, sheet):
        self.wb = wb
        self.sheet = sheet

    @staticmethod
    def load(path, sheetName=None, dataOnly=False):
        import openpyxl as exc
        wb = exc.load_workbook(filename=path, data_only=dataOnly)
        if sheetName:
            sheet = wb[sheetName]
        else:
            sheet = wb.active
        return OpenpyxlSheet(wb, sheet)

    def getNumberOfRows(self) -> int:
        return self.sheet.max_row

    def getNumberOfCols(self) -> int:
        return self.sheet.max_column

    def value(self, col: int, row: int):
        cell = self.sheet.cell(column=col+1, row=row+1)
        if cell:
            v = cell.value
            if isinstance(v, str):
                return v.strip()
            else:
                return v

    def setValue(self, col: int, row: int, value: any):
        self.sheet.cell(column=col+1, row=row+1, value=value)

    def save(self, path):
        saveBook(self.wb, path)

# Xlwings supports formulae but only works on Windows and Mac. And it is slow.
# So only use if absolutely necessary
def xlCoord(col: int, row: int):
    return (row+1, col+1)

_xlApp = None

def closeXlApp():
    if _xlApp:
        _xlApp.quit()

def xlApp():
    global _xlApp
    if _xlApp is not None:
        return _xlApp
    import xlwings as xw
    atexit.register(closeXlApp)
    _xlApp = xw.App(visible=False)

    return _xlApp

MAX_COL = 1000
MAX_ROW = 1000

class XlwingsSheet(Sheet):
    def __init__(self, wb, sheet):
        self.wb = wb
        self.sheet = sheet

    @staticmethod
    def load(path, sheetName=None, dataOnly=False):
        app = xlApp()
        verbose(f'Loading xlwings book from {path}')
        wb = app.books.open(path)
        if sheetName:
            sheet = wb.sheets[sheetName]
        else:
            sheet = wb.sheets[0]
        return XlwingsSheet(wb, sheet)

    def getNumberOfRows(self) -> int:
        return 1000

    def getNumberOfCols(self) -> int:
        return 1000

    def value(self, col: int, row: int):
        v = self.sheet.range(xlCoord(col=col, row=row)).value
        if isinstance(v, str):
            return v.strip()
        else:
            return v

    def setValue(self, col: int, row: int, value: any):
        r = self.sheet.range(xlCoord(col=col, row=row))
        r.value = value

    def save(self, path):
        saveBook(self.wb, path)

    def close(self):
        verbose('closing xlwings book')
        self.wb.close()
