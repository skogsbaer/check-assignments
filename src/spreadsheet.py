import os
import tempfile
from typing import *
from ownLogging import *

def saveExcelSpreadsheet(path: str, wb):
    # I experienced problem with corrupt .xlsx files, probably when the python process
    # was interrupted while writing the data.
    with tempfile.NamedTemporaryFile(suffix='.xlsx', dir=os.path.dirname(path)) as fp:
        fp.close()
        wb.save(fp.name)
        os.rename(fp.name, path) # Atomic move

def openSheet(path: str, sheetName: Optional[str], dataOnly=False):
    import openpyxl as exc
    wb = exc.load_workbook(filename=path, data_only=dataOnly)
    if sheetName:
        sheet = wb[sheetName]
    else:
        sheet = wb.active
    return (wb, sheet)

def cellValue(sheet, column, row):
    cell = sheet.cell(column=column, row=row)
    if cell and cell.value:
        v = cell.value
        if isinstance(v, str):
            return v.strip()
        else:
            return v
    else:
        return None

def findColumnIndex(sheet, title: Union[str, list[str]]) -> list[int]:
    if not isinstance(title, list):
        title = [title]
    ixs = []
    for col in range(1, sheet.max_column + 1):
        v = cellValue(sheet, column=col, row=1)
        if v in title:
            ixs.append(col)
    return ixs

def findRowIndex(sheet, title: Union[str, list[str]]) -> list[int]:
    if not isinstance(title, list):
        title = [title]
    ixs = []
    for row in range(1, sheet.max_row + 1):
        v = cellValue(sheet, column=1, row=row)
        if v in title:
            ixs.append(row)
    return ixs

def getDataFromRow(path: str, sheetName: str, rowTitle: Union[str, list[str]]):
    (_wb, sheet) = openSheet(path, sheetName, dataOnly=True)
    ixs = findRowIndex(sheet, rowTitle)
    if len(ixs) != 1:
        raise ValueError("Sheet {sheetName} of spreadsheet {path} has more than one row with " \
            f"titles {rowTitle}")
    row = ixs[0]
    for col in range(2, sheet.max_column + 1):
        v = cellValue(sheet, column=col, row=row)
        if v:
            return v
    return None

def replaceData(path: str, colTitle: str, oldValue: str, newValue: str, sheetName: Optional[str] = None):
    (wb, sheet) = openSheet(path, sheetName)
    idxs = findColumnIndex(sheet, colTitle)
    if not idxs:
        raise ValueError(f"No column found with title {colTitle}")
    if len(idxs) > 1:
        raise ValueError(f"More than one column found with title {colTitle}")
    idx = idxs[0]
    for row in range(1, sheet.max_row + 1):
        v = cellValue(sheet, column=idx, row=row)
        if v:
            if v == oldValue:
                sheet.cell(column=idx, row=row, value=newValue)
    saveExcelSpreadsheet(path, wb)

def enterData(path: str,
    idColumn: str | list[str], idValue: Union[str, list[str]],
    dataColumn: str, data: str,
    sheetName: Optional[str] = None):
    """
    Enters some data into the spreadsheet stored at path.
    - idColumn is the name of the column containing IDs for the rows. This column must exist.
    - idValue is the value that is searched in idColum
    - dataColumn is the name of the column used for storing the data. This column may or may not exist.
      The name of the column might also be splitted in two rows
    - data is the data to be stored.
    """
    if not isinstance(idValue, list):
        idValue = [idValue]
    (wb, sheet) = openSheet(path, sheetName)
    idColumnIxs = findColumnIndex(sheet, idColumn)
    if not idColumnIxs:
        raise ValueError(f'No column named {idColumn} found at {path} ({sheetName})')
    dataColumnIxs = findColumnIndex(sheet, dataColumn)
    if not dataColumnIxs:
        # Check whether the name of the dataColumn is split over two rows
        last = ''
        firstRow = []
        for col in range(1, sheet.max_column + 1):
            v = cellValue(sheet, column=col, row=1)
            if v:
                v = str(v)
                last = v
            else:
                v = last
            firstRow.append(v)
        for col in range(1, sheet.max_column + 1):
            v = cellValue(sheet, column=col, row=2)
            if v:
                v = str(v)
                if col - 1 < len(firstRow):
                    oneAbove = firstRow[col - 1]
                    if (oneAbove + ' ' + v) == dataColumn or (oneAbove + v) == dataColumn:
                        if dataColumnIxs:
                            raise ValueError(f'Duplicate dataColumn {dataColumn} at {path}')
                        dataColumnIxs = [col]
    if not dataColumnIxs:
        dataColumnIx = sheet.max_column + 1
        sheet.cell(column=dataColumnIx, row=1, value=dataColumn)
    else:
        dataColumnIx = dataColumnIxs[0]
    rowIx = None
    for row in range(1, sheet.max_row + 1):
        for idColumnIx in idColumnIxs:
            v = cellValue(sheet, column=idColumnIx, row=row)
            if v:
                if v in idValue:
                    if rowIx is not None and rowIx != row:
                        raise ValueError(f'Duplicate ID {idValue} in column {idColumn} at {path}')
                    rowIx = row
    if rowIx is None:
        raise ValueError(f'No row found with value {idValue} in column {idColumn} at {path}')
    verbose(f'Storing {data} at column={dataColumnIx}, row={rowIx} in {path} (sheet: {sheetName})')
    sheet.cell(column=dataColumnIx, row=rowIx, value=data)
    saveExcelSpreadsheet(path, wb)
